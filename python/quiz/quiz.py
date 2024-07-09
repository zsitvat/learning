import pandas as pd
import tkinter as tk
from tkinter import scrolledtext, filedialog
import os
from openpyxl import load_workbook


def load_excel(file_path):
    """Load an Excel file and return a DataFrame with columns: Question, Answer, and Question type."""
    try:
        df = pd.read_excel(file_path)
        if df.empty:
            raise ValueError("The selected file contains no questions.")

        # Normalize True/False values and replace Hungarian values with English equivalents
        if "True/False" in df.columns:
            df["True/False"] = df["True/False"].astype(str).str.strip().str.capitalize()
            df["True/False"] = df["True/False"].replace({"Igaz": "True", "Hamis": "False"})

        # Initialize a new DataFrame with the required columns
        data = {
            "Question": [],
            "Answer": [],
            "Question type": []
        }

        if "Question type" in df.columns:
            data["Question"].extend(df["Question"])
            data["Answer"].extend(df["Answer"])
            data["Question type"].extend(df["Question type"])

        else:

            # Process True/False questions
            if "Statements" in df.columns and "True/False" in df.columns:
                true_false_questions = df[df["True/False"].notna()]
                data["Question"].extend(true_false_questions["Statements"])
                data["Answer"].extend(true_false_questions["True/False"])
                data["Question type"].extend(["True/False"] * len(true_false_questions))

            # Process Fill-in-the-blank questions
            if "Sentences" in df.columns and "Words" in df.columns:
                fill_in_the_blank_questions = df[df["Words"].notna()]
                data["Question"].extend(fill_in_the_blank_questions["Sentences"])
                data["Answer"].extend(fill_in_the_blank_questions["Words"])
                data["Question type"].extend(["Fill-in-the-blank"] * len(fill_in_the_blank_questions))

        # Create a DataFrame from the processed data
        final_df = pd.DataFrame(data)

        print(final_df)

        if final_df.empty:
            raise ValueError("The file must have either True/False or Fill-in-the-blank questions with proper columns.")

        return final_df
    except Exception as e:
        return str(e)



class MainApp:
    """Main application class for the True or False quiz."""

    def __init__(self, root):
        self.root = root
        self.df = pd.DataFrame(columns=["Question", "Answer", "Question type"])
        self.num_questions = 0
        self.error_message = None
        self.choice = None
        self.data_frame = None

        self.root.title("Quiz Application")
        self.root.geometry("700x500+650+300")

        self.root.protocol("WM_DELETE_WINDOW", lambda: self.close_program())

        self.show_file_dialog()

    def show_file_dialog(self):
        """Show the initial file dialog to choose the file or search for a file."""
        for widget in self.root.winfo_children():
            widget.destroy()

        label = tk.Label(self.root, text="Which file would you like to open?", font=("Arial", 12))
        label.pack(pady=10)

        self.error_label = tk.Label(self.root, text="", font=("Arial", 12), fg="red")
        self.error_label.pack(pady=5)

        button_frame = tk.Frame(self.root)
        button_frame.pack(pady=10)

        incorrect_button = tk.Button(
            button_frame, text="Incorrect", font=("Arial", 12), command=lambda: self.on_choice("incorrect")
        )
        incorrect_button.pack(side="left", padx=5)

        correct_button = tk.Button(
            button_frame, text="Correct", font=("Arial", 12), command=lambda: self.on_choice("correct")
        )
        correct_button.pack(side="left", padx=5)

        search_button = tk.Button(
            button_frame, text="Search File", font=("Arial", 12), command=self.search_file
        )
        search_button.pack(side="left", padx=5)

    def on_choice(self, choice):
        """Handle the choice of correct or incorrect questions."""
        data_frame = load_excel(f"quiz/{choice}_questions.xlsx")
        if isinstance(data_frame, str):
            self.show_error(data_frame)
        else:
            self.choice = choice
            self.data_frame = data_frame
            self.show_num_questions_dialog()

    def search_file(self):
        """Open a file dialog to search for an Excel file."""
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            data_frame = load_excel(file_path)
            if isinstance(data_frame, str):
                self.show_error(data_frame)
            else:
                self.choice = "custom"
                self.data_frame = data_frame
                self.show_num_questions_dialog()

    def show_error(self, message):
        """Display an error message on the screen."""
        self.error_label.config(text=message)

    def show_num_questions_dialog(self):
        """Show the dialog to choose the number of questions for the quiz."""
        if self.choice:
            data_frame = self.data_frame

            if data_frame.empty:
                CustomMessageBox(self.root, "Error", "The selected file contains no questions. Exiting.")
                self.root.mainloop()
                self.root.destroy()
                return

            max_length = len(data_frame)
            dialog = NumQuestionsDialog(self.root, max_length)
            self.root.wait_window(dialog)

            if dialog.value:
                self.num_questions = dialog.value
                self.df = data_frame
                self.root.withdraw()
                quiz_window = tk.Toplevel(self.root)
                QuizApp(quiz_window, self.num_questions, self.df, self.root)
            else:
                self.close_program()

    def close_program(self):
        """Close the program."""
        self.root.quit()
        self.root.destroy()


class QuizApp:
    """Class for the quiz application."""

    def __init__(self, window, num_questions, data_frame, main_root):
        self.window = window
        self.df = data_frame
        self.window.title("Quiz Application")
        self.window.geometry("700x500+650+300")
        self.num_questions = num_questions
        self.main_root = main_root
        self.window.protocol("WM_DELETE_WINDOW", lambda: self.close_program(self.window))
        self.create_widgets()
        self.reset_quiz()

    def reset_quiz(self):
        """Reset the quiz to start again."""
        self.questions = self.df.sample(n=self.num_questions).reset_index(drop=True)
        self.current_question = 0
        self.score = 0
        self.answers = []
        self.show_question()

    def create_widgets(self):
        """Create the widgets for the quiz application."""
        self.statement_label = tk.Label(self.window, wraplength=600, justify="center", font=("Arial", 18))
        self.statement_label.pack(pady=20)

        self.button_frame = tk.Frame(self.window)
        self.button_frame.pack(pady=10)

        self.true_button = tk.Button(self.button_frame, text="True", font=("Arial", 12), width=12, command=lambda: self.check_answer("True"))
        self.true_button.pack(side="left", padx=20)
        self.true_button.bind("<Return>", lambda event: self.check_answer("True"))

        self.false_button = tk.Button(self.button_frame, text="False", font=("Arial", 12), width=12, command=lambda: self.check_answer("False"))
        self.false_button.pack(side="right", padx=20)
        self.false_button.bind("<Return>", lambda event: self.check_answer("False"))

        self.fill_in_entry = tk.Entry(self.window, font=("Arial", 12), width=30)
        self.fill_in_entry.pack(pady=20)
        self.fill_in_entry.pack_forget()

        self.submit_button = tk.Button(self.window, text="Submit", font=("Arial", 12), command=self.submit_fill_in_answer)
        self.submit_button.pack(pady=10)
        self.submit_button.pack_forget()
        self.submit_button.bind("<Return>", self.submit_fill_in_answer)

        self.confirmation_label = tk.Label(self.window, text="", font=("Arial", 12))
        self.confirmation_label.pack(pady=10)

    def show_question(self):
        """Display the current question based on its type."""
        if self.current_question < len(self.questions):
            question = self.questions.iloc[self.current_question]
            question_type = question["Question type"]
            if question_type == "True/False":
                self.statement_label.config(text=question["Question"])
                self.true_button.pack(side="left", padx=20)
                self.false_button.pack(side="right", padx=20)
                self.fill_in_entry.pack_forget()
                self.submit_button.pack_forget()
                self.fill_in_entry.unbind("<Return>")
            elif question_type == "Fill-in-the-blank":
                self.statement_label.config(text=question["Question"])
                self.true_button.pack_forget()
                self.false_button.pack_forget()
                self.fill_in_entry.pack(pady=20)
                self.submit_button.pack(pady=10)
                self.fill_in_entry.bind("<Return>", self.submit_fill_in_answer)
        else:
            self.show_result()

    def submit_fill_in_answer(self, event=None):
        """Submit the answer for fill-in-the-blank questions."""
        self.check_answer()

    def check_answer(self, answer=None):
        """Validate the answer based on Question type."""
        question = self.questions.iloc[self.current_question]
        question_type = question["Question type"]
        correct_answer = str(question["Answer"]).strip().lower()
        if question_type == "True/False":
            user_answer = str(answer).strip().lower()
        elif question_type == "Fill-in-the-blank":
            user_answer = self.fill_in_entry.get().strip().lower()
            self.fill_in_entry.delete(0, tk.END)  # Reset text input after submission

        self.answers.append((question["Question"], user_answer, correct_answer))
        if "vagy" in correct_answer:
            correct_answer = correct_answer.split(" vagy ")
            if user_answer in correct_answer:
                self.score += 1
                self.show_info("Correct!", "Your answer is correct!", correct=True)
            else:
                self.show_info("Incorrect", f"Wrong answer! The correct answer is [{correct_answer}]", correct=False)
        else:
            
            if user_answer == correct_answer:
                self.score += 1
                self.show_info("Correct!", "Your answer is correct!", correct=True)
            else:
                self.show_info("Incorrect", f"Wrong answer! The correct answer is [{correct_answer}]", correct=False)

        self.current_question += 1
        if self.current_question < len(self.questions):
            self.show_question()
        else:
            self.show_result()

    def show_info(self, title, message, correct):
        """Display an information dialog."""
        info_dialog = CustomMessageBox(self.window, title, message, correct)
        info_dialog.geometry("400x200+850+400")
        self.window.wait_window(info_dialog)

    def show_result(self):
        """Display the quiz results."""
        self.window.withdraw()
        percent_score = (self.score / len(self.questions)) * 100
        result_window = tk.Toplevel(self.window)
        result_window.geometry("800x600+600+300")
        result_window.title("Quiz Results")
        result_window.protocol("WM_DELETE_WINDOW", lambda: self.close_program(result_window))

        result_label = tk.Label(result_window, text=f"Your score: {self.score}/{len(self.questions)}\nPercentage: {percent_score:.2f}%", font=("Arial", 16))
        result_label.pack(pady=10)

        text_area = scrolledtext.ScrolledText(result_window, wrap=tk.WORD, width=80, height=20, font=("Arial", 12))
        text_area.pack(pady=10, padx=10)

        text_area.tag_config("correct", foreground="green")
        text_area.tag_config("incorrect", foreground="red")

        for statement, user_answer, correct_answer in self.answers:
            text_area.insert(tk.END, f"Question: {statement}\n")
            if user_answer.lower() == correct_answer.lower():
                text_area.insert(tk.END, f"Your Answer: {user_answer}\n", "correct")
            else:
                text_area.insert(tk.END, f"Your Answer: {user_answer}\n", "incorrect")
            text_area.insert(tk.END, f"Correct Answer: {correct_answer}\n\n")

        text_area.config(state=tk.DISABLED)

        button_frame = tk.Frame(result_window)
        button_frame.pack(pady=10)

        save_correct_button = tk.Button(
            button_frame, text="Save Correct", font=("Arial", 12), command=self.save_correct
        )
        save_correct_button.pack(side="left", padx=5)
        save_correct_button.bind("<Return>", self.save_correct)

        save_incorrect_button = tk.Button(
            button_frame, text="Save Incorrect", font=("Arial", 12), command=self.save_incorrect
        )
        save_incorrect_button.pack(side="left", padx=5)
        save_incorrect_button.bind("<Return>", self.save_incorrect)

        close_button = tk.Button(
            button_frame, text="Close", font=("Arial", 12), command=lambda: self.close_program(result_window)
        )
        close_button.pack(side="left", padx=5)
        close_button.bind("<Return>", lambda event: self.close_program(result_window))

        restart_button = tk.Button(
            button_frame, text="Start Again", font=("Arial", 12), command=lambda: self.restart_quiz(result_window)
        )
        restart_button.pack(side="right", padx=5)
        restart_button.bind("<Return>", lambda event: self.restart_quiz(result_window))

        self.confirmation_label_result = tk.Label(result_window, text="", font=("Arial", 12))
        self.confirmation_label_result.pack(pady=10)

    def save_correct(self, event=None):
        """Save the correct questions to an Excel file."""
        file_path = "quiz/correct_questions.xlsx"

        correct_questions = [
            (ans[0], ans[2], "True/False" if ans[2] in ["True", "False", "true", "false"] else "Fill-in-the-blank")
            for ans in self.answers
            if ans[1].lower() == ans[2].lower()
        ]

        df_new_correct = pd.DataFrame(correct_questions, columns=["Question", "Answer", "Question type"])

        if not os.path.exists(file_path):
            # If the file does not exist, create it and write the data
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                if not df_new_correct.empty:
                    df_new_correct.to_excel(writer, index=False, sheet_name="Correct Questions")
        else:
            # If the file exists, open it and append the new data
            with pd.ExcelWriter(file_path, engine="openpyxl", mode='a', if_sheet_exists='overlay') as writer:
                if not df_new_correct.empty:
                    df_existing = pd.read_excel(file_path, sheet_name="Correct Questions")
                    df_combined = pd.concat([df_existing, df_new_correct], ignore_index=True)
                    df_combined.to_excel(writer, index=False, sheet_name="Correct Questions")

        # Load the workbook to ensure sheet visibility
        wb = load_workbook(file_path)
        if "Correct Questions" in wb.sheetnames:
            wb["Correct Questions"].sheet_state = 'visible'

        # Ensure at least one sheet is visible
        visible_sheets = [sheet for sheet in wb.sheetnames if wb[sheet].sheet_state == 'visible']
        if not visible_sheets:
            wb.active.sheet_state = 'visible'

        wb.save(file_path)

        self.confirmation_label_result.config(
            text="Correct questions have been saved to correct_questions.xlsx"
        )

    def save_incorrect(self, event=None):
        """Save the incorrect questions to an Excel file."""
        file_path = "quiz/incorrect_questions.xlsx"

        incorrect_questions = [
            (ans[0], ans[2], "True/False" if ans[2] in ["True", "False", "true", "false"] else "Fill-in-the-blank")
            for ans in self.answers
            if ans[1].lower() != ans[2].lower()
        ]

        df_new_incorrect = pd.DataFrame(incorrect_questions, columns=["Question", "Answer", "Question type"])

        if not os.path.exists(file_path):
            # If the file does not exist, create it and write the data
            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                if not df_new_incorrect.empty:
                    df_new_incorrect.to_excel(writer, index=False, sheet_name="Incorrect Questions")
        else:
            # If the file exists, open it and append the new data
            with pd.ExcelWriter(file_path, engine="openpyxl", mode='a', if_sheet_exists='overlay') as writer:
                if not df_new_incorrect.empty:
                    df_existing = pd.read_excel(file_path, sheet_name="Incorrect Questions")
                    df_combined = pd.concat([df_existing, df_new_incorrect], ignore_index=True)
                    df_combined.to_excel(writer, index=False, sheet_name="Incorrect Questions")

        # Load the workbook to ensure sheet visibility
        wb = load_workbook(file_path)
        if "Incorrect Questions" in wb.sheetnames:
            wb["Incorrect Questions"].sheet_state = 'visible'

        # Ensure at least one sheet is visible
        visible_sheets = [sheet for sheet in wb.sheetnames if wb[sheet].sheet_state == 'visible']
        if not visible_sheets:
            wb.active.sheet_state = 'visible'

        wb.save(file_path)

        self.confirmation_label_result.config(
            text="Incorrect questions have been saved to incorrect_questions.xlsx"
        )

    def close_program(self, window):
        """Close the program."""
        window.destroy()
        self.main_root.deiconify()

    def restart_quiz(self, window):
        """Restart the quiz."""
        window.destroy()
        self.main_root.deiconify()
        MainApp(self.main_root)


class NumQuestionsDialog(tk.Toplevel):
    """Dialog window to choose the number of questions for the quiz."""

    def __init__(self, parent, max_questions, error_message=None):
        super().__init__(parent)
        self.parent = parent
        self.max_questions = max_questions
        self.value = None
        self.geometry("400x200+780+400")
        self.title("Number of Questions")

        label = tk.Label(
            self, text="How many questions would you like to answer?", font=("Arial", 12)
        )
        label.pack(pady=10)
        label = tk.Label(self, text=f"Number of questions: {self.max_questions}", font=("Arial", 12))
        label.pack(pady=10)

        self.entry = tk.Entry(self, font=("Arial", 12))
        self.entry.pack(pady=5)
        self.entry.bind("<Return>", self.on_ok)  # Bind Enter key to OK button

        self.error_label = tk.Label(self, text="", font=("Arial", 12), fg="red")
        self.error_label.pack(pady=5)

        button_frame = tk.Frame(self)
        button_frame.pack(pady=10)

        ok_button = tk.Button(button_frame, text="OK", font=("Arial", 12), command=self.on_ok)
        ok_button.pack(side="left", padx=5)

        cancel_button = tk.Button(button_frame, text="Cancel", font=("Arial", 12), command=self.on_cancel)
        cancel_button.pack(side="right", padx=5)

        if error_message:
            self.show_error(error_message)

    def on_ok(self, event=None):
        """Handle the OK button click."""
        try:
            value = int(self.entry.get())
            if 1 <= value <= self.max_questions:
                self.value = value
                self.destroy()
            else:
                self.show_error(f"Please enter a number between 1 and {self.max_questions}.")
        except ValueError:
            self.show_error("Please enter a valid number.")

    def on_cancel(self):
        """Handle the Cancel button click."""
        self.value = None
        self.destroy()

    def show_error(self, message):
        """Display an error message on the screen."""
        self.error_label.config(text=message)


class CustomMessageBox(tk.Toplevel):
    """Custom message box to display messages."""

    def __init__(self, parent, title, message, correct=None):
        super().__init__(parent)
        self.geometry("400x200+850+400")
        self.title(title)
        bg_color = "green" if correct else "red" if correct is not None else None
        self.config(bg=bg_color)

        label = tk.Label(self, text=message, font=("Arial", 12), bg=bg_color)
        label.pack(pady=10)

        button = tk.Button(self, text="OK", font=("Arial", 12), command=self.on_ok)
        button.pack(pady=10)
        button.bind("<Return>", self.on_ok)  # Bind Enter key to OK button

    def on_ok(self, event=None):
        """Handle the OK button click."""
        self.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = MainApp(root)
    root.mainloop()
